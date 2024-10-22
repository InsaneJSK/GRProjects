import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']

for i in range(2, sheet.max_row+1):
    sheet[f'c{i}'].value *= 0.9

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length (mm)'

data = Reference(sheet, min_col=3, min_row=1, max_row=sheet.max_row, max_col=3)
chart1.add_data(data, titles_from_data=True)
chart1.shape = 4
sheet.add_chart(chart1, f"A{sheet.max_row+3}")

wb.save('transactions-updt.xlsx')
